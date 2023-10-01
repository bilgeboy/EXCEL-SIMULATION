from flask import jsonify
from openpyxl import load_workbook
import pandas as pd
import json
import io


def ret_row(file, file_name):
    # try:
    #     df = pd.read_excel(file)
    # except Exception as e:
    #     return f'Error reading XLSX file: {str(e)}', 500
    #
    #     # Convert the DataFrame to a list of dictionaries (JSON-like structure)
    # data = df.to_dict(orient='records')
    #
    # # Return the data as JSON response
    # return jsonify(data)

    # Load the Excel file
    path = file_name
    df = pd.read_excel(path, engine='xlrd')
    path = file_name.replace(".xls", ".xlsx")
    df.to_excel(path, index=False)
    workbook = load_workbook(path,
                             data_only=True)  # Use data_only=True to get cell values instead of formulas

    # Create an empty dictionary to store sheet data
    all_data = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Create a list to store data for the current sheet
        sheet_data = []

        row_data = []

        # Iterate through rows in the sheet and store data in a list of dictionaries
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                row_data.append(cell)
            sheet_data.append(row_data)
            row_data = []

        # Store the sheet data in the dictionary using the sheet name as the key
        all_data[sheet_name] = sheet_data

    # Print or return the JSON data
    return all_data

